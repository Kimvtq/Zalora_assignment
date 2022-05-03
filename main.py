import requests
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl
import time

# define a method to create a new folder and direct to it
def create_then_go_to_folder(folder_name):
    folder_dir = os.path.join(os.getcwd(), folder_name)
    try:
        os.mkdir(folder_dir)
    except:
        pass
    os.chdir(folder_dir)

# define a method to get to detailed page of a SKU then download all images
def download_image(sku):
    # Direct to search result page for a sku
    driver.get("https://www.zalora.com.my/catalog/?q=" + sku)

    
    try:
        # Get the result and click on the link
        result_link = driver.find_element(By.XPATH,'//li[@id="' + sku + '"]//a')
        result_link.click()

        # Get all thumbnails except video 
        thumbnails  = driver.find_elements(By.XPATH,'//ul[contains(@class,"prd-moreImagesList")]//li[not(div)]')

        for thumbmail in thumbnails:
            # thumbmail.click()
        
            # img = driver.find_element(By.CSS_SELECTOR,'.prd-image')
            name = thumbmail.get_attribute("data-image-alt")
            link = thumbmail.get_attribute('data-image-big')
            with open(name.replace(' ', '-').replace('/', '') + '.jpg', 'wb') as f:
                im = requests.get(link)
                f.write(im.content)
    except:
        pass

# Initial driver
driver = webdriver.Chrome("ChromeDriver\chromedriver.exe")   
driver.maximize_window()

# Location of the file
path = "sku_list.xlsx"

# to open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
 
max_row = sheet_obj.max_row

create_then_go_to_folder("media")
media_dir = os.getcwd()

# Loop will print all values
# of first column
for i in range(2, max_row + 1):
    cell_obj = sheet_obj.cell(row = i, column = 1)
    sku = cell_obj.value
    if sku != None:
        print(sku) # to check sku 
        sku.strip();
        create_then_go_to_folder(sku)
        download_image(sku)
        # go back to media folder
        os.chdir(media_dir)


# time.sleep(10)




